from os import fspath
import xlwings as excel
import xlrd
import openpyxl
import json

#ghi vào file excel mình muốn ( tạo mới )
class write_excel:
    def __init__(self,links_excel) -> None:
        self.links_excel = links_excel
        self.app = excel.App(visible = False)
        self.workbook = self.app.books.add()

    def write(self,sheet_name,data_want_to_write):
        self.workbook.sheets[0].name = sheet_name
        wsheet = self.workbook.sheets[0]
        last_row = wsheet.range(f"A{wsheet.cells.last_cell.row}").end("up").row + 1
        yellow = (255, 255, 0)
        rowheader = wsheet.range("A1:D1")
        rowheader.color = yellow
        rowheader.api.Font.Bold = True
        rowheader.api.Borders.Weight = 2

        wsheet.range(f"A{last_row}").value = data_want_to_write

        wsheet.used_range.api.Borders.Weight = 2
        wsheet.used_range.api.WrapText = False
        wsheet.autofit()

    """tắt file excel"""
    def save(self):
        
        self.workbook.save(self.links_excel)
        self.workbook.close()
        self.app.quit()
        print("______________________Save Complete_______________________")


# class Read_excel:
#     def __init__(self,links_excel) -> None:
#         self.links_excel = links_excel
#         self.app = excel.App(visible = False)
#         self.workbook = self.app.books.open(self.links_excel )


#     def Read_Define(self):
#         wsheet = self.workbook.sheets["Define"]
#         list_define_meaning = wsheet.range("A:A").value
#         list_define_name = wsheet.range("B:B").value
#         list_define_value= wsheet.range("C:C").value
#         list_define_file = wsheet.range("D:D").value
#         return list_define_meaning, list_define_name, list_define_value, list_define_file

#     """tắt file excel"""
#     def close(self):
#         self.workbook.close()
#         self.app.quit()
#         print("______________________Close Complete_______________________")

###end class read excel 1

class Read_excel:
    def __init__(self,links_excel) -> None:
        self.xl_workbook = xlrd.open_workbook(links_excel)
        self.sheet_define = self.xl_workbook.sheet_by_name("Define")
        # self.data_use_range = [[self.sheet_define.cell_value(r, c) for c in range(self.sheet_define.ncols)] for r in range(self.sheet_define.nrows)] # all data in sheet


    def Read_Define(self):
        sheet = self.sheet_define
        list_define_meaning = [sheet.cell_value(r, 0)  for r in range(sheet.nrows)]
        list_define_name  = [sheet.cell_value(r, 1)  for r in range(sheet.nrows)]
        list_define_value = [sheet.cell_value(r, 2)  for r in range(sheet.nrows)]
        list_define_file  = [sheet.cell_value(r, 3)  for r in range(sheet.nrows)]
        print("______________________Read Define Complete_______________________")
        return list_define_meaning, list_define_name, list_define_value, list_define_file

    def save_database_json(self):
        define_meaning, define_name, define_value, define_file = self.Read_Define()
        Dic_database = {}
        Dic_database["Define_meaning"] = define_meaning
        Dic_database["Define_name"] = define_name
        Dic_database["Define_value"] = define_value
        Dic_database["Define_file"] = define_file
        save_file_json(".\\Database\\Database.json",Dic_database)

    def read_database_json(self):
        database = read_file_json(".\\Database\\Database.json")
        return database


##### end class read excel 2


# class Read_excel:
#     def __init__(self,links_excel) -> None:
#         self.link = links_excel
#         self.workbook = openpyxl.load_workbook(self.link , data_only=True)
#         self.sheet_data = self.workbook["Define"]
#         self.max_row = self.sheet_data.max_row
#         self.max_column = self.sheet_data.max_column

#     def Read_Define(self):
#         list_define_meaning = []
#         list_define_name    = []
#         list_define_value   = []
#         list_define_file    = []

#         for i in range(self.max_row):
#             string_meaning = self.sheet_data.cell(row=i+1,column=1).value
#             string_name = self.sheet_data.cell(row=i+1,column=2).value
#             string_value = self.sheet_data.cell(row=i+1,column=3).value
#             string_file = self.sheet_data.cell(row=i+1,column=4).value
#             list_define_meaning.append(string_meaning if string_meaning != None else "")
#             list_define_name.append(string_name if string_name != None else "")
#             list_define_value.append(string_value if string_value != None else "")
#             list_define_file.append(string_file if string_file != None else "")
#         print("______________________Read Define Complete_______________________")
#         return list_define_meaning, list_define_name, list_define_value, list_define_file

##### end class read excel 3

def save_file_json(path_save,dic_data):
    # Write JSON file
    with open(path_save, 'w', encoding='utf-8', errors="ignore") as outfile:
        json.dump(dic_data,outfile,indent=4, ensure_ascii=False)
        outfile.close()
    return path_save

def read_file_json(path_json):
    # Write JSON file
    with open(path_json, 'r', encoding='utf-8', errors="ignore") as outfile:
        json_data = json.load(outfile)
        outfile.close()
    return json_data

if __name__ == '__main__':
    # exxcel = write_excel("Database.xlsx")
    # exxcel.write("Define","")
    # exxcel = Read_excel("Database.xlsx")
    # list1, list2, list3, list4 = exxcel.Read_Define()
    # for data in list1:
    #     print(data)
    # for data in list4:
    #     print(data)
    exxcel = Read_excel("Database.xlsx")
    exxcel.save_database_json()
    exxcel.read_database_json()