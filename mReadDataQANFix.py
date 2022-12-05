'''
Created on May 17, 2022

@author: nhanbuivan
'''
import openpyxl
import os
import re
import json
import Class_read_links_file
import Class_write_excel
class DataQA:
    
    data = {    "Enum" : {},
                "Struct" : {},
                "Global" : {},
                "Const" : {},
                "Macro" : {},
            }
    
    def __init__(self,linkExcel):
        self.workbook = openpyxl.load_workbook(linkExcel , data_only=True)
        self.sheetNames = self.workbook.sheetnames
        self.checkData = ["_Enumeration", "_Structure", "_Global_variable", "_Constant", "_Macro", "_Local_variable"]
        self.dataType = ["Enum", "Struct", "Global", "Const", "Macro"]
        
    def addData(self, sheetname, dataType):
        sheet = self.workbook[sheetname]
        ## đọc hàng đầu tiên đẻ bắt dữ liệu
        rowMax = sheet.max_row  
        rowStart = 0
        for i in range(rowMax):
            string = str(sheet.cell(row=i+1,column=1).value).lower().strip()
            if string == "no.":
                rowStart = i + 1
                break
        ## bắt đầu lọc dữ liệu
        for i in range(rowStart, rowMax):
            nameType = sheet.cell(i + 1, 4).value
            dataType1 = sheet.cell(i + 1, 5).value
            dataType2 = sheet.cell(i + 1, 6).value
            dataType3 = sheet.cell(i + 1, 7).value
            
            if nameType not in DataQA.data[dataType]:
                DataQA.data[dataType][nameType] = []
                
            if dataType == self.dataType[0]:  ## cho kiểu enum
                ## dataType1 nếu  = "-" thì bỏ qua, nếu khác "-" nó là 1 phần tử con của kiểu enum đó
                ## dataType2 là í nghĩa của kiểu enum nếu dataType1 = "-", ngược lại thì là í nghĩa của kiểu sub enum
                data = [dataType1, dataType2]
                DataQA.data[dataType][nameType].append(data.copy())
            elif dataType == self.dataType[1]: ## cho kiểu struct
                ## dataType1 nếu  = "-" thì bỏ qua, nếu khác "-" nó là 1 phần tử con của kiểu struct đó
                ## dataType2 là í nghĩa của kiểu struct nếu dataType1 = "-", ngược lại thì là í nghĩa của kiểu sub struct
                data = [dataType1, dataType2]
                DataQA.data[dataType][nameType].append(data.copy())
            elif dataType == self.dataType[2]: ## cho variable global
                ## dataType1: range của biến
                ## dataType2: là lsn của biến
                ## dataType3: là í nghĩa của biến
                data = [dataType1, dataType2, dataType3]
                DataQA.data[dataType][nameType] =data.copy()
            elif dataType == self.dataType[3]: ## cho kiểu const
                ## dataType1: í nghĩa của const
                data = [dataType1]
                DataQA.data[dataType][nameType] =data.copy()
            elif dataType == self.dataType[4]: ## cho macro
                ## dataType1: í nghĩa của macro
                ## dataType2: là lsn của macro
                data = [dataType1, dataType2]
                DataQA.data[dataType][nameType] =data.copy()
            
    def exportData(self):
        for nameSheet in self.sheetNames:
            if nameSheet.endswith(self.checkData[0]):
                self.addData(nameSheet, self.dataType[0])
            elif nameSheet.endswith(self.checkData[1]):
                self.addData(nameSheet, self.dataType[1])
            elif nameSheet.endswith(self.checkData[2]):
                self.addData(nameSheet, self.dataType[2])
            elif nameSheet.endswith(self.checkData[3]):
                self.addData(nameSheet, self.dataType[3]) 
            elif nameSheet.endswith(self.checkData[4]):
                self.addData(nameSheet, self.dataType[4])
     
     
def fixDatabase_InoutDD_FormQA(dic, dicFix, type = "struct"):
    for mainObject in dic:
        if mainObject == None:
            continue
        if type == "const":
            if mainObject in dicFix:
                dicFix[mainObject][0] = dic[mainObject][0]
            elif (mainObject + "[]") in dicFix:
                dicFix[(mainObject + "[]")][0] = dic[mainObject][0]
            else:
                print(mainObject + " Not exits in " + "Database_DDIO_" + type)
        elif type == "variable":
            if mainObject in dicFix:
                dicFix[mainObject][0] = dic[mainObject][2] #### í nghĩa
                dicFix[mainObject][3] = dic[mainObject][0]  #### range của biến
                dicFix[mainObject][4] = dic[mainObject][1]  #### lsb của biến
            elif (mainObject + "[]") in dicFix:
                dicFix[(mainObject + "[]")][0] = dic[mainObject][2] #### í nghĩa
                dicFix[(mainObject + "[]")][3] = dic[mainObject][0]  #### range của biến
                dicFix[(mainObject + "[]")][4] = dic[mainObject][1]  #### lsb của biến
            else:
                print(mainObject + " Not exits in " + "Database_DDIO_" + type)
        else:
            for elements in dic[mainObject]:
                if type == "struct" or type == "enum":
                    if elements[0].strip() == "-":
                        if mainObject in dicFix:
                            dicFix[mainObject][0] = elements[1]  #### meaning of object - in file name - "-"(don't care)
                        else:
                            print(mainObject + " Not exits in " + "Database_DDIO_" + type)
                    else:
                        if elements[0].strip() in dicFix:
                            dicFix[elements[0].strip()][0] = elements[1] #### meaning of object - in file name - type of object
                        else:
                            print(elements[0].strip() + " Not exits in " + "Database_DDIO_" + type)
                            
    pathSaveDatabase = ".\\Database\\Database_DDIO_" + type + ".json"
    Class_write_excel.save_file_json(pathSaveDatabase,dicFix)
               
def fixDatabaseINOUT_FromQA(linkFolder):
    allFile = Class_read_links_file.Export_file(linkFolder)
    allFile.pattern_file = ".+\.xlsx" # xuất tất cả các file excel
    ############ đọc dữ liệu từ QA
    for file in allFile.List_All_file:
        xxx = DataQA(file)
        xxx.exportData()
        
 
    dataFileStruct = Class_write_excel.read_file_json(".\\Database\\Database_DDIO_struct.json")
    dataFileEnum = Class_write_excel.read_file_json(".\\Database\\Database_DDIO_enum.json")
    dataFileVariable = Class_write_excel.read_file_json(".\\Database\\Database_DDIO_variable.json")
    dataFileConst = Class_write_excel.read_file_json(".\\Database\\Database_DDIO_const.json")
    
    fixDatabase_InoutDD_FormQA(DataQA.data[xxx.dataType[1]], dataFileStruct, "struct")
    fixDatabase_InoutDD_FormQA(DataQA.data[xxx.dataType[0]], dataFileEnum, "enum")
    fixDatabase_InoutDD_FormQA(DataQA.data[xxx.dataType[2]], dataFileVariable, "variable")
    fixDatabase_InoutDD_FormQA(DataQA.data[xxx.dataType[3]], dataFileConst, "const")
    
    print("Complete !!!")
               
if __name__ == "__main__":
    
    # pathFileExcel = readLinkFile()
    # xxx = DataQA(pathFileExcel)
    # xxx.exportData()
    # for _ in DataQA.data[xxx.dataType[4]]:
    #     print(_, ":", DataQA.data[xxx.dataType[4]][_])

    # def save_file_json(path_save,dic_data):
    #     # Write JSON file
    #     with open(path_save, 'w', encoding='utf-8', errors="ignore") as outfile:
    #         json.dump(dic_data,outfile,indent=4, ensure_ascii=False)
    #         outfile.close()
    #     return path_save
    #
    # def read_file_json(path_json):
    #     # Write JSON file
    #     with open(path_json, 'r', encoding='utf-8', errors="ignore") as outfile:
    #         json_data = json.load(outfile)
    #         outfile.close()
    #     return json_data
    #
    # allFile = Class_read_links_file.Export_file(r"C:\Projects\Schaeffler_Phase_6\01_Working\05_Output_DD\01_QA_DD")
    # allFile.pattern_file = ".+\.xlsx" # xuất tất cả các file excel
    # ############ đọc dữ liệu từ QA
    # for _ in allFile.List_All_file:
    #     xxx = DataQA(_)
    #     xxx.exportData()
    #
    # # for _ in DataQA.data[xxx.dataType[3]]:
    # #     print(_, ":", DataQA.data[xxx.dataType[3]][_])
    #
    # # print(len(DataQA.data[xxx.dataType[0]]))
    # ############ đọc dữ liệu từ QA########################### end
    #
    #
    # dataFileStruct = read_file_json(".\\Database\\Database_DDIO_struct.json")
    # dataFileEnum = read_file_json(".\\Database\\Database_DDIO_enum.json")
    # dataFileVariable = read_file_json(".\\Database\\Database_DDIO_variable.json")
    # dataFileConst = read_file_json(".\\Database\\Database_DDIO_const.json")
    #
    # fixDatabase_InoutDD_FormQA(DataQA.data[xxx.dataType[1]], dataFileStruct, "struct")
    # fixDatabase_InoutDD_FormQA(DataQA.data[xxx.dataType[0]], dataFileEnum, "enum")
    # fixDatabase_InoutDD_FormQA(DataQA.data[xxx.dataType[2]], dataFileVariable, "variable")
    # fixDatabase_InoutDD_FormQA(DataQA.data[xxx.dataType[3]], dataFileConst, "const")
    # print("Complete !!!")

    fixDatabaseINOUT_FromQA(r"C:\Projects\Schaeffler_Phase_6\01_Working\05_Output_DD\01_QA_DD")













